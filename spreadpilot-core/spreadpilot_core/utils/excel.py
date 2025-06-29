"""Excel report generation utilities for SpreadPilot."""

import datetime
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..logging import get_logger
from ..models import Follower
from .time import format_ny_time

logger = get_logger(__name__)


def generate_excel_report(
    output_path: str,
    follower: Follower,
    month: int,
    year: int,
    pnl_total: float,
    commission_amount: float,
    daily_pnl: dict[str, float],
) -> str:
    """Generate Excel report for a follower.

    Args:
        output_path: Output directory path
        follower: Follower model
        month: Report month (1-12)
        year: Report year
        pnl_total: Total P&L for the month
        commission_amount: Commission amount
        daily_pnl: Dict mapping dates (YYYYMMDD) to daily P&L

    Returns:
        Path to generated Excel file
    """
    try:
        # Create output directory if it doesn't exist
        os.makedirs(output_path, exist_ok=True)

        # Generate filename
        filename = f"spreadpilot-report-{year}-{month:02d}-{follower.id}.xlsx"
        filepath = os.path.join(output_path, filename)

        # Create workbook
        wb = openpyxl.Workbook()

        # Get active sheet
        ws = wb.active
        ws.title = "Monthly Report"

        # Set column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

        # Define styles
        title_font = Font(name="Arial", size=14, bold=True)
        header_font = Font(name="Arial", size=12, bold=True)
        normal_font = Font(name="Arial", size=11)

        header_fill = PatternFill(
            start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Add title
        month_name = datetime.date(year, month, 1).strftime("%B")
        ws["A1"] = f"SpreadPilot Monthly Report - {month_name} {year}"
        ws["A1"].font = title_font
        ws.merge_cells("A1:D1")
        ws["A1"].alignment = Alignment(horizontal="center")

        # Add follower information
        ws["A3"] = "Follower Information"
        ws["A3"].font = header_font
        ws.merge_cells("A3:D3")

        # Follower info headers
        ws["A4"] = "ID"
        ws["B4"] = follower.id
        ws["A5"] = "Email"
        ws["B5"] = follower.email
        ws["A6"] = "IBAN"
        ws["B6"] = follower.iban
        ws["A7"] = "Commission Rate"
        ws["B7"] = f"{follower.commission_pct}%"

        # Style follower info
        for row in range(4, 8):
            ws[f"A{row}"].font = normal_font
            ws[f"A{row}"].border = thin_border
            ws[f"A{row}"].fill = header_fill
            ws[f"B{row}"].font = normal_font
            ws[f"B{row}"].border = thin_border

        # Add monthly summary
        ws["A9"] = "Monthly Summary"
        ws["A9"].font = header_font
        ws.merge_cells("A9:D9")

        # Summary headers
        ws["A10"] = "Total P&L"
        ws["B10"] = pnl_total
        ws["B10"].number_format = "$#,##0.00"
        ws["A11"] = "Commission Rate"
        ws["B11"] = f"{follower.commission_pct}%"
        ws["A12"] = "Commission Amount"
        ws["B12"] = commission_amount
        ws["B12"].number_format = "$#,##0.00"

        # Style summary
        for row in range(10, 13):
            ws[f"A{row}"].font = normal_font
            ws[f"A{row}"].border = thin_border
            ws[f"A{row}"].fill = header_fill
            ws[f"B{row}"].font = normal_font
            ws[f"B{row}"].border = thin_border

        # Add daily P&L table
        ws["A14"] = "Daily P&L"
        ws["A14"].font = header_font
        ws.merge_cells("A14:D14")

        # Daily P&L headers
        ws["A15"] = "Date"
        ws["B15"] = "P&L"

        # Style headers
        ws["A15"].font = normal_font
        ws["A15"].border = thin_border
        ws["A15"].fill = header_fill
        ws["B15"].font = normal_font
        ws["B15"].border = thin_border
        ws["B15"].fill = header_fill

        # Sort dates
        sorted_dates = sorted(daily_pnl.keys())

        # Add daily P&L data
        for i, date in enumerate(sorted_dates):
            row = 16 + i
            # Format date from YYYYMMDD to YYYY-MM-DD
            formatted_date = f"{date[:4]}-{date[4:6]}-{date[6:]}"
            ws[f"A{row}"] = formatted_date
            ws[f"B{row}"] = daily_pnl[date]
            ws[f"B{row}"].number_format = "$#,##0.00"

            # Style cells
            ws[f"A{row}"].font = normal_font
            ws[f"A{row}"].border = thin_border
            ws[f"B{row}"].font = normal_font
            ws[f"B{row}"].border = thin_border

        # Add footer
        footer_row = 16 + len(sorted_dates) + 2
        ws[f"A{footer_row}"] = (
            f"This report was generated on {format_ny_time()} by SpreadPilot."
        )
        ws[f"A{footer_row + 1}"] = (
            "For any questions, please contact capital@tradeautomation.it."
        )

        ws.merge_cells(f"A{footer_row}:D{footer_row}")
        ws.merge_cells(f"A{footer_row + 1}:D{footer_row + 1}")

        # Save workbook
        wb.save(filepath)

        logger.info(
            "Generated Excel report",
            follower_id=follower.id,
            month=month,
            year=year,
            filepath=filepath,
        )

        return filepath
    except Exception as e:
        logger.error(
            f"Error generating Excel report: {e}",
            follower_id=follower.id,
            month=month,
            year=year,
        )
        raise
